from openpyxl import load_workbook
import re
import pandas as pd
from os import getenv
import tempfile

pattern = r"FT\d+-\d+"


async def parse_xlsx(file):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
        print(tmp_path)

        # Check the size of the temporary file
        # tmp.seek(0, 2)
        # tmp_size = tmp.tell()
        # print(f"Temporary file size: {tmp_size} bytes")

        # tmp.seek(0)
        workbook = load_workbook(
            filename=tmp_path,
            read_only=True,
        )

        cell_vals = set()
        for wsh in workbook.sheetnames:
            wbk = workbook[wsh]
            for row in wbk.iter_rows():
                for cell in row:
                    match = re.match(pattern, str(cell.value))
                    if match:
                        cell_vals.add(match.group(0))

        df = pd.DataFrame(
            {
                "Batch_ID": list(cell_vals),
                "Filename": [file.filename for _ in range(len(cell_vals))],
            },
        )
        # print(df)
        return df


def parse_generic(file, cro, project):
    lst = []
    doc, num, suffix = file.filename.split("-")
    batch = f"{num}-{suffix.split('.')[0]}"
    lst.append([num, batch, file.filename, doc, cro, project])
    df = pd.DataFrame(
        lst,
        columns=[
            "Compound_ID",
            "Batch_ID",
            "Filename",
            "Doc_type",
            "CRO",
            "Project",
        ],
    )

    print(df)
    return df


def parse_other_file_types(file):
    lst = []
    matches = re.findall(pattern, file.filename)
    if matches:
        for match in matches:
            lst.append([match, file.filename])
    df = pd.DataFrame(
        lst,
        columns=[
            "Batch_ID",
            "Filename",
        ],
    )

    print(df)
    return df
